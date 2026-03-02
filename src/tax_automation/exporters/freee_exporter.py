"""freee取引インポート用Excelエクスポーター"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from ..models import PaymentMethod, Transaction


# freee取引インポートの列定義
_FREEE_COLUMNS = [
    "収支区分",
    "管理番号",
    "発生日",
    "決済期日",
    "取引先",
    "勘定科目",
    "税区分",
    "金額",
    "税計算区分",
    "税額",
    "備考",
    "品目",
    "部門",
    "メモタグ",
    "セグメント1",
    "セグメント2",
    "セグメント3",
    "決済口座",
    "決済日",
    "決済方法",       # 追加: カード/現金
    "購入品目",       # 追加: レシートからの品目詳細
]

# 勘定科目ごとのデフォルト税区分（飲食店向け）
_TAX_CATEGORY_MAP = {
    "仕入高": "課対仕入10%",
    "消耗品費": "課対仕入10%",
    "被服費": "課対仕入10%",
    "通信費": "課対仕入10%",
    "交通費": "課対仕入10%",
    "会議費": "課対仕入10%",
    "接待交際費": "課対仕入10%",
    "広告宣伝費": "課対仕入10%",
    "外注費": "課対仕入10%",
    "新聞図書費": "課対仕入10%",
    "地代家賃": "課対仕入10%",
    "水道光熱費": "課対仕入10%",
    "福利厚生費": "課対仕入10%",
    "損害保険料": "対象外",
    "修繕費": "課対仕入10%",
    "雑費": "課対仕入10%",
}


class FreeeExcelExporter:
    """Transaction リストをfreee取引インポート用Excelに変換して出力する。"""

    def __init__(self, settlement_account: str = "ごうぎんVISAカード"):
        self.settlement_account = settlement_account

    def export(
        self,
        transactions: list[Transaction],
        output_path: Path | str,
        year: int | None = None,
    ) -> Path:
        """トランザクションをfreee用Excelに書き出す。"""
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "取引データ"

        # ヘッダー行
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(_FREEE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # データ行
        row_idx = 2
        for tx in sorted(transactions, key=lambda t: t.date):
            if year and tx.date.year != year:
                continue

            amount = int(tx.amount)
            # マイナス金額（返金）は収入として処理
            if amount < 0:
                income_expense = "収入"
                amount = abs(amount)
            else:
                income_expense = "支出"

            category_name = tx.category_name or tx.category_code or "雑費"
            tax_category = _TAX_CATEGORY_MAP.get(tx.category_code or "", "課対仕入10%")

            # 決済方法による決済口座の切り替え
            is_cash = tx.payment_method == PaymentMethod.CASH
            settlement_account = "現金" if is_cash else self.settlement_account
            payment_method_label = "現金" if is_cash else "カード"

            # レシート品目の整形
            items_text = ""
            if tx.receipt_items:
                items_text = ", ".join(
                    f"{item.name}(\xa5{int(item.amount):,})" if item.amount else item.name
                    for item in tx.receipt_items[:10]
                )
                if len(tx.receipt_items) > 10:
                    items_text += f" 他{len(tx.receipt_items) - 10}品"

            # 備考欄: Gemini判断理由 or レシート品目の要約
            remarks = tx.gemini_reasoning if tx.gemini_reasoning else ""

            row_data = [
                income_expense,                          # 収支区分
                "",                                      # 管理番号
                tx.date.strftime("%Y/%m/%d"),            # 発生日
                "",                                      # 決済期日
                tx.merchant_name,                        # 取引先
                category_name,                           # 勘定科目
                tax_category,                            # 税区分
                amount,                                  # 金額
                "税込",                                  # 税計算区分
                "",                                      # 税額（自動計算）
                remarks,                                 # 備考
                "",                                      # 品目
                "",                                      # 部門
                "",                                      # メモタグ
                "",                                      # セグメント1
                "",                                      # セグメント2
                "",                                      # セグメント3
                settlement_account,                      # 決済口座
                tx.date.strftime("%Y/%m/%d"),            # 決済日
                payment_method_label,                    # 決済方法
                items_text,                              # 購入品目
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx == 8:  # 金額列
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")

            row_idx += 1

        # 列幅の調整
        column_widths = {
            1: 10, 2: 10, 3: 12, 4: 12, 5: 30,
            6: 18, 7: 15, 8: 12, 9: 12, 10: 10,
            11: 40, 12: 10, 13: 10, 14: 10, 15: 12,
            16: 12, 17: 12, 18: 22, 19: 12, 20: 10, 21: 50,
        }
        for col, width in column_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        wb.save(path)
        return path
